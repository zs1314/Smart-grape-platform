import os
import sys
import cv2
import numpy as np
import openpyxl
import pandas as pd
import psutil
import pynvml
from PyQt5.QtCore import Qt, QTimer, QDateTime, QUrl
from PyQt5.QtGui import QCursor, QPixmap
from PyQt5.QtWidgets import QApplication, QMainWindow, QDesktopWidget, QFileDialog, QInputDialog, QTableWidgetItem, \
    QMessageBox
from window1 import *
from window2 import *
from window3 import *
from window4 import *
from window5 import *
from my_mobileNet_nn import demo
from my_ui_yolov5_ import detect


# 窗口一
class my_window1(QMainWindow):
    def __init__(self):
        super(my_window1, self).__init__()
        self.ui_1 = Ui_window_1()
        self.ui_1.setupUi(self)
        # 消除边框
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        # 阴影
        self.shadow = QtWidgets.QGraphicsDropShadowEffect(self)
        self.shadow.setOffset(0, 0)
        self.shadow.setBlurRadius(15)
        self.shadow.setColor(QtCore.Qt.black)
        self.ui_1.frame.setGraphicsEffect(self.shadow)
        # 按钮与信号相联系
        self.ui_1.pushButton_3.clicked.connect(self.on_pushButton_close)
        self.ui_1.pushButton_close.clicked.connect(self.on_pushButton_min)
        self.ui_1.pushButton_jichushezhi.clicked.connect(self.on_button_jichushezhi)
        self.ui_1.pushButton_wenshicanshu.clicked.connect(self.on_button_wenshicanshu)
        self.ui_1.pushButton_putaofenlei.clicked.connect(self.on_button_putaofenlei)
        self.ui_1.pushButton_shishijiance.clicked.connect(self.on_button_shishijiance)
        self.ui_1.pushButton_duquwenjian.clicked.connect(self.on_button_duquwenjian)
        self.ui_1.pushButton_pintaijianjie.clicked.connect(self.on_button_pintaijianjie)
        self.ui_1.pushButton_wenshigo.clicked.connect(self.wenshi)
        self.ui_1.pushButton_25.clicked.connect(self.putaofenlei)
        self.ui_1.pushButton_31.clicked.connect(self.shihsijiance)
        self.ui_1.pushButton_32.clicked.connect(self.wenjianchuli)
        # 定时器
        timer_gpu = QTimer(self)  # 定义一个定时器对象
        timer_gpu.timeout.connect(self.lookgpu)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer_gpu.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

        timer_cpu = QTimer(self)  # 定义一个定时器对象
        timer_cpu.timeout.connect(self.lookcpu)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer_cpu.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

        timer_time = QTimer(self)  # 定义一个定时器对象
        timer_time.timeout.connect(self.showtime)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer_time.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

        timer_button_selected = QTimer(self)  # 定义一个定时器对象
        timer_button_selected.timeout.connect(self.button_selected)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer_button_selected.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

    def button_selected(self):
        index = self.ui_1.stackedWidget.currentIndex()
        if index == 0:
            self.ui_1.pushButton_jichushezhi.setStyleSheet("background-color:  #2b2b2b;")
        else:
            self.ui_1.pushButton_jichushezhi.setStyleSheet("background-color:#676767;")
        if index == 1:
            self.ui_1.pushButton_wenshicanshu.setStyleSheet("background-color:  #2b2b2b;")
        else:
            self.ui_1.pushButton_wenshicanshu.setStyleSheet("background-color:#676767;")
        if index == 2:
            self.ui_1.pushButton_putaofenlei.setStyleSheet("background-color:  #2b2b2b;")
        else:
            self.ui_1.pushButton_putaofenlei.setStyleSheet("background-color:#676767;")
        if index == 3:
            self.ui_1.pushButton_shishijiance.setStyleSheet("background-color:  #2b2b2b;")
        else:
            self.ui_1.pushButton_shishijiance.setStyleSheet("background-color:#676767;")
        if index == 4:
            self.ui_1.pushButton_duquwenjian.setStyleSheet("background-color:  #2b2b2b;")
        else:
            self.ui_1.pushButton_duquwenjian.setStyleSheet("background-color:#676767;")
        if index == 5:
            self.ui_1.pushButton_pintaijianjie.setStyleSheet("background-color:  #2b2b2b;")
        else:
            self.ui_1.pushButton_pintaijianjie.setStyleSheet("background-color:#676767;")

    def wenshitishi(self):
        msg_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, '提示',
                                        '已连接数据库')  # Information可替换为Warning、Critical其他提示框类型
        msg_box.setWindowIcon(QtGui.QIcon('img_icon/提示_logo.png'))  # 加载图标
        msg_box.exec_()

    def wenshi(self):
        win1.hide()
        win2.show()
        self.wenshitishi()

    def putaofenlei(self):
        win1.hide()
        win3.show()

    def shihsijiance(self):
        win1.hide()
        win4.show()

    def wenjianchuli(self):
        win1.hide()
        win5.show()

    """调整窗口位置"""

    # 窗口居中
    def center(self):
        screen = QDesktopWidget().screenGeometry()
        size = self.geometry()
        self.move((screen.width() - size.width()) / 2,
                  (screen.height() - size.height()) / 2)

    """无边框控制移动"""

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()  # 获取鼠标相对窗口的位置
            event.accept()
            self.setCursor(QCursor(Qt.OpenHandCursor))  # 更改鼠标图标

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_flag:
            self.move(QMouseEvent.globalPos() - self.m_Position)  # 更改窗口位置
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False
        self.setCursor(QCursor(Qt.ArrowCursor))

    def on_pushButton_min(self):
        # 最小化
        self.showMinimized()

    def on_pushButton_close(self):
        # 关闭程序
        # 后两项分别为按钮(以|隔开，共有7种按钮类型，见示例后)、默认按钮(省略则默认为第一个按钮)
        reply = QMessageBox.warning(self, "警告", "是否关闭程序", QMessageBox.Yes | QMessageBox.No,
                                    QMessageBox.Yes)
        # msg_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, '警告',
        #                                 '是否关闭程序', QMessageBox.Yes | QMessageBox.No,
        #                                 QMessageBox.Yes)  # Information可替换为Warning、Critical其他提示框类型
        # msg_box.setWindowIcon(QtGui.QIcon('img_icon/警告_logo.png'))  # 加载图标
        # msg_box.exec_()
        if reply==16384:
            self.close()

    def on_button_jichushezhi(self):
        self.ui_1.stackedWidget.setCurrentIndex(0)

    def on_button_wenshicanshu(self):
        self.ui_1.stackedWidget.setCurrentIndex(1)

    def on_button_putaofenlei(self):
        self.ui_1.stackedWidget.setCurrentIndex(2)

    def on_button_shishijiance(self):
        self.ui_1.stackedWidget.setCurrentIndex(3)

    def on_button_duquwenjian(self):
        self.ui_1.stackedWidget.setCurrentIndex(4)

    def on_button_pintaijianjie(self):
        self.ui_1.stackedWidget.setCurrentIndex(5)

    def showtime(self):
        # 展示系统时间
        datetime = QDateTime.currentDateTime()  # 获取系统时间
        text = datetime.toString()
        self.ui_1.label_time.setText("   " + text)  # 在空间上展示

    def lookgpu(self):
        server_info_list = []
        UNIT = 1024 * 1024
        pynvml.nvmlInit()  # 初始化
        gpu_device_count = pynvml.nvmlDeviceGetCount()  # 获取Nvidia GPU块数
        for gpu_index in range(gpu_device_count):
            handle = pynvml.nvmlDeviceGetHandleByIndex(gpu_index)  # 获取GPU i的handle，后续通过handle来处理
            memery_info = pynvml.nvmlDeviceGetMemoryInfo(handle)  # 通过handle获取GPU 的信息
            server_info_list.append(
                {
                    "gpu_id": gpu_index,  # gpu id
                    "total": int(memery_info.total / UNIT),  # gpu 总内存
                    "used": int(memery_info.used / UNIT),  # gpu使用内存
                    "utilization": pynvml.nvmlDeviceGetUtilizationRates(handle).gpu  # 使用率
                }
            )
            gpu_name = str(pynvml.nvmlDeviceGetName(handle))
            gpu_temperature = pynvml.nvmlDeviceGetTemperature(handle, 0)
            gpu_power_state = pynvml.nvmlDeviceGetPowerState(handle)
            gpu_util_rate = pynvml.nvmlDeviceGetUtilizationRates(handle).gpu
            gpu_memory_rate = pynvml.nvmlDeviceGetUtilizationRates(handle).memory
            # print(f"第 %d 张卡：{gpu_index}")
            # print(f"显卡名：{gpu_name}")
            # print(f"内存总容量：{memery_info.total / UNIT} MB")
            # print(f"使用容量：{memery_info.total / UNIT}MB")
            # print(f"剩余容量：{memery_info.total / UNIT}MB")
            # print(f"显存空闲率：{memery_info.free / memery_info.total}")
            # print(f"温度：{gpu_temperature}摄氏度")
            # print(f"供电水平：{gpu_power_state}")
            # print(f"gpu计算核心满速使用率：{gpu_util_rate}")
            # print(f"gpu内存读写满速使用率：{gpu_memory_rate}")
            # print(f"内存占用率：{memery_info.used / memery_info.total}")
            text = str(gpu_util_rate) + '%'
            self.ui_1.label_GPU.setText(text)
        pynvml.nvmlShutdown()  # 关闭管理工具

    def lookcpu(self):
        """
        获取CPU使用率
        :return:  使用率，（%无小数）
        """
        cpu_percent = psutil.cpu_percent(interval=0)
        cpu_info = "%.0f%%" % cpu_percent
        self.ui_1.label_CPU.setText(cpu_info)


# 窗口二
class my_window2(QMainWindow):
    def __init__(self):
        super(my_window2, self).__init__()
        self.ui_2 = Ui_window_2()
        self.ui_2.setupUi(self)
        # 消除边框
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.shadow = QtWidgets.QGraphicsDropShadowEffect(self)
        self.shadow.setOffset(0, 0)
        self.shadow.setBlurRadius(18)
        self.shadow.setColor(QtCore.Qt.black)
        self.ui_2.frame.setGraphicsEffect(self.shadow)
        # 按钮与信号
        self.ui_2.pushButton_min.clicked.connect(self.on_pushButton_min)
        self.ui_2.pushButton_close.clicked.connect(self.on_pushButton_close)
        self.ui_2.pushButton_back.clicked.connect(self.back)

        # 计时器
        timer_time = QTimer(self)  # 定义一个定时器对象
        timer_time.timeout.connect(self.showtime)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer_time.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

        self.database()

    def showtime(self):
        # 展示系统时间
        datetime = QDateTime.currentDateTime()  # 获取系统时间
        text = datetime.toString()
        self.ui_2.label_time.setText("   " + text)  # 在空间上展示

    def back(self):
        win2.hide()
        win1.show()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()  # 获取鼠标相对窗口的位置
            event.accept()
            self.setCursor(QCursor(Qt.OpenHandCursor))  # 更改鼠标图标

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_flag:
            self.move(QMouseEvent.globalPos() - self.m_Position)  # 更改窗口位置
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False
        self.setCursor(QCursor(Qt.ArrowCursor))

    def on_pushButton_min(self):
        # 最小化
        self.showMinimized()

    def on_pushButton_close(self):
        # 关闭程序
        self.close()

    def database(self):
        import pymysql

        # 打开数据库连接
        # 连接名称、用户民、密码、数据库名（需已经存在的）
        db = pymysql.connect(host="localhost", user="root", password="123456", db="conservatory")

        # 使用 cursor() 方法创建一个游标对象 cursor
        cursor = db.cursor()
        sql1 = 'select * from arg'
        # 使用 execute()  方法执行 SQL 查询
        cursor.execute(sql1)
        # 使用 fetchone() 方法获取单条数据.
        data = cursor.fetchall()
        for i in data:
            print(i)
        self.ui_2.textBrowser.setText(str(i[0]) + '°C')
        self.ui_2.textBrowser_2.setText(str(i[1]) + '%')
        self.ui_2.textBrowser_3.setText(str(i[2]))
        self.ui_2.textBrowser_4.setText(str(i[3]))
        self.ui_2.textBrowser_5.setText(str(i[4]) + 'lx')
        self.ui_2.textBrowser_6.setText(str(i[5]) + '%')
        # 关闭游标
        cursor.close()
        # 关闭数据库连接
        db.close()


# 窗口三
class my_window3(QMainWindow):
    def __init__(self):
        super(my_window3, self).__init__()
        self.ui_3 = Ui_window_3()
        self.ui_3.setupUi(self)
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.shadow = QtWidgets.QGraphicsDropShadowEffect(self)
        self.shadow.setOffset(0, 0)
        self.shadow.setBlurRadius(30)
        self.shadow.setColor(QtCore.Qt.blue)
        self.ui_3.frame.setGraphicsEffect(self.shadow)
        self.labels = ['欧塞瓦', '品丽珠', '赤霞珠', '霞多丽', '梅洛', '米勒图高', '黑皮诺', '雷司令', '长相思 ',
                       '西拉', '丹魄']
        self.ui_3.pushButton_close.clicked.connect(self.on_pushButton_close)
        self.ui_3.pushButton_min.clicked.connect(self.on_pushButton_min)
        self.ui_3.pushButton_back.clicked.connect(self.back)
        self.ui_3.pushButton.clicked.connect(self.showimage)
        self.ui_3.stopButton.clicked.connect(self.stoptest)
        self.ui_3.pushButton_4.clicked.connect(self.save)
        # 定时器
        timer_time = QTimer(self)  # 定义一个定时器对象
        timer_time.timeout.connect(self.showtime)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer_time.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

        self.timer_shipin = QtCore.QTimer()  # 创建一个定时器对象
        self.timer_shipin.timeout.connect(self.show_shipin)  # 定时器对象.timeout指的是当start()结束时，所调用的函数（进行操作）
        self.ui_3.pushButton_2.clicked.connect(self.shipin_button)  # 摄像头按钮绑定video_button这个槽函数

        # 摄像头的定时器
        self.timer_camera = QtCore.QTimer()  # 创建一个定时器对象
        self.timer_camera.timeout.connect(self.show_camera)  # 定时器对象.timeout指的是当start()结束时，所调用的函数（进行操作）
        self.ui_3.pushButton_3.clicked.connect(self.video_button)  # 摄像头按钮绑定video_button这个槽函数
        # 媒体播放
        # self.player=QMediaPlayer(self)
        # 媒体播放定时器
        # self.timer_video=QtCore.QTimer()
        # self.timer_video.timeout.connect(self.changeSlide)
        # # self.player.positionChanged.connect(self.get_time)
        # self.player.positionChanged.connect(self.changeSlide)
        # 标志
        self.flag_img = 0
        self.flag_shipin = 0
        self.flag_camera = 0

        self.ui_3.checkBox.stateChanged.connect(self.camtishi)

    def showtime(self):
        # 展示系统时间
        datetime = QDateTime.currentDateTime()  # 获取系统时间
        text = datetime.toString()
        self.ui_3.label_time.setText("   " + text)  # 在空间上展示

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()  # 获取鼠标相对窗口的位置
            event.accept()
            self.setCursor(QCursor(Qt.OpenHandCursor))  # 更改鼠标图标

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_flag:
            self.move(QMouseEvent.globalPos() - self.m_Position)  # 更改窗口位置
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False
        self.setCursor(QCursor(Qt.ArrowCursor))

    def on_pushButton_min(self):
        # 最小化
        self.showMinimized()

    def on_pushButton_close(self):
        # 关闭程序
        self.close()

    def back(self):
        win3.hide()
        win1.show()

    """----------弹窗提示-----------"""

    def messageDialog(self):
        # 弹窗核心功能代码
        msg_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Critical, '警告',
                                        '请先暂停检测')  # Information可替换为Warning、Critical其他提示框类型
        msg_box.setWindowIcon(QtGui.QIcon('img_icon/警告_logo.png'))  # 加载图标
        msg_box.exec_()

    def messageDialog_wenjian(self):
        msg_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, '警告',
                                        '未选择文件')  # Information可替换为Warning、Critical其他提示框类型
        msg_box.setWindowIcon(QtGui.QIcon('img_icon/警告_logo.png'))  # 加载图标
        msg_box.exec_()

    def camtishi(self):
        msg_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, '提示',
                                        '已改变cam状态')  # Information可替换为Warning、Critical其他提示框类型
        msg_box.setWindowIcon(QtGui.QIcon('img_icon/提示_logo.png'))  # 加载图标
        msg_box.exec_()

    def save_messageDialog(self, path):
        msg_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, '提示',
                                        '已保存结果至' + path)  # Information可替换为Warning、Critical其他提示框类型
        msg_box.setWindowIcon(QtGui.QIcon('img_icon/提示_logo.png'))  # 加载图标
        msg_box.exec_()

    """----------弹窗提示-----------"""

    """------------图片检测------------"""

    def showimage(self):
        if not self.ui_3.checkBox.isChecked():
            print(self.ui_3.checkBox.isChecked())
            if self.flag_img == 0:

                """原图(未处理)"""
                global imgNamepath  # 这里为了方便别的地方引用图片路径，将其设置为全局变量
                # 弹出一个文件选择框，第一个返回值imgName记录选中的文件路径+文件名，第二个返回值imgType记录文件的类型
                # QFileDialog就是系统对话框的那个类第一个参数是上下文，第二个参数是弹框的名字，第三个参数是默认打开的路径，第四个参数是需要的格式
                imgNamepath, imgType = QFileDialog.getOpenFileName(self, "选择图片",
                                                                   "",
                                                                   "*.jpg;;*.png;;All Files(*)")
                print(imgType)

                if len(imgNamepath) > 0:
                    self.flag_img = 1
                    # 先保证是否能打开
                    # 通过文件路径获取图片文件，并设置图片长宽为label控件的长、宽
                    my_img = QPixmap(imgNamepath).scaled(self.ui_3.label_8.width(), self.ui_3.label_8.height())
                    print(imgNamepath)
                    # 在label控件上显示选择的图片
                    self.ui_3.label_8.setPixmap(my_img)

                    """处理后"""
                    shrink_process, pred_idx, confs = demo.predict_img(imgNamepath)  # 返回的numppy格式,正好是opencv
                    # shrink_process = cv2.cvtColor(process_img, cv2.COLOR_BGR2RGB)  # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的
                    # 图像处理 通用  针对RGB三通道图片显示
                    QtImg_process = QtGui.QImage(shrink_process.data,
                                                 shrink_process.shape[1],
                                                 shrink_process.shape[0],
                                                 shrink_process.shape[1] * 3,
                                                 QtGui.QImage.Format_RGB888)

                    # 修改图片的高宽,适应控件的大小
                    jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
                        self.ui_3.label_9.width(), self.ui_3.label_9.height())

                    # 在控件上放置图片
                    self.ui_3.label_9.setPixmap(jpg_out_process)
                    for i in range(5):
                        pred_class = self.labels[pred_idx[i]]  # 类别
                        text = str(pred_class) + str('  ') + str(confs[i])  # 打印出来的格式
                        self.ui_3.listWidget.addItem(text)
                else:
                    self.messageDialog_wenjian()
            else:
                self.messageDialog()
        else:
            if self.flag_img == 0:

                """原图(未处理)"""
                # global imgNamepath  # 这里为了方便别的地方引用图片路径，将其设置为全局变量
                # 弹出一个文件选择框，第一个返回值imgName记录选中的文件路径+文件名，第二个返回值imgType记录文件的类型
                # QFileDialog就是系统对话框的那个类第一个参数是上下文，第二个参数是弹框的名字，第三个参数是默认打开的路径，第四个参数是需要的格式
                imgNamepath, imgType = QFileDialog.getOpenFileName(self, "选择图片",
                                                                   "",
                                                                   "*.jpg;;*.png;;All Files(*)")
                print(imgType)

                if len(imgNamepath) > 0:
                    self.flag_img = 1
                    # 先保证是否能打开
                    # 通过文件路径获取图片文件，并设置图片长宽为label控件的长、宽
                    my_img = QPixmap(imgNamepath).scaled(self.ui_3.label_8.width(), self.ui_3.label_8.height())
                    print(imgNamepath)
                    # 在label控件上显示选择的图片
                    self.ui_3.label_8.setPixmap(my_img)

                    """处理后"""
                    shrink_process, pred_idx, confs = demo.cam(imgNamepath)  # 返回的numppy格式,正好是opencv
                    # shrink_process = cv2.cvtColor(process_img, cv2.COLOR_BGR2RGB)  #
                    # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的 图像处理 通用  针对RGB三通道图片显示
                    QtImg_process = QtGui.QImage(shrink_process.data,
                                                 shrink_process.shape[1],
                                                 shrink_process.shape[0],
                                                 shrink_process.shape[1] * 3,
                                                 QtGui.QImage.Format_RGB888)

                    # 修改图片的高宽,适应控件的大小
                    jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
                        self.ui_3.label_9.width(), self.ui_3.label_9.height())

                    # 在控件上放置图片
                    self.ui_3.label_9.setPixmap(jpg_out_process)
                    for i in range(5):
                        pred_class = self.labels[pred_idx[i]]  # 类别
                        text = str(pred_class) + str('  ') + str(confs[i])  # 打印出来的格式
                        self.ui_3.listWidget.addItem(text)
                else:
                    self.messageDialog_wenjian()
            else:
                self.messageDialog()

    """----------图片检测-----------"""

    """----------视频检测-----------"""

    def shipin_button(self):
        if self.flag_shipin == 0:

            # 弹出一个文件选择框，第一个返回值imgName记录选中的文件路径+文件名，第二个返回值imgType记录文件的类型
            # QFileDialog就是系统对话框的那个类第一个参数是上下文，第二个参数是弹框的名字，第三个参数是默认打开的路径，第四个参数是需要的格式
            shipinNamepath, shipinType = QFileDialog.getOpenFileName(self, "选择视频",
                                                                     "",
                                                                     "*.mp4;;*.avi;;All Files(*)")
            # print(shipinType)
            if len(shipinNamepath) > 0:
                self.flag_shipin = 1
                # 得到文件后缀名  需要根据情况进行修改
                suffix = shipinNamepath.split("/")[-1][shipinNamepath.split("/")[-1].index(".") + 1:]
                print(shipinNamepath, suffix)
                # self.player.setMedia(QMediaContent(QUrl.fromLocalFile(shipinNamepath)))
                if suffix == "mp4" or suffix == "avi":
                    self.cap_shipin = cv2.VideoCapture(shipinNamepath)  # 0表示摄像头，返回值就是整个视屏资源
                    self.timer_shipin.start()
            else:
                self.messageDialog_wenjian()
        else:
            self.messageDialog()

    def show_shipin(self):
        ret, self.shipin_img = self.cap_shipin.read()  # 读取每一帧图片，read（）返回两个参数：1、是否读取成功（布尔值） 2、这一帧的图片
        if ret:
            # 若成功读取——就展示（调用下面这个方法）
            self.show_shipin_img(self.shipin_img)

    # def changeSlide(self,position):
    #     self.vidoeLength = self.player.duration() + 0.1
    #     print(self.vidoeLength)
    #     self.ui_3.progressBar.setValue(round((position / self.vidoeLength) * 100))
    #     # self.ui_3.progressBar.setValue(round(self.player.position() * self.maxValue / self.player.duration()))

    def show_shipin_img(self, img):
        if not self.ui_3.checkBox.isChecked():
            """处理前"""
            self.ui_3.listWidget.clear()
            """对返回图片进行处理"""
            shrink = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg = QtGui.QImage(shrink.data,
                                 shrink.shape[1],
                                 shrink.shape[0],
                                 shrink.shape[1] * 3,
                                 QtGui.QImage.Format_RGB888)
            # print(QtImg)
            # 修改图片的高宽
            jpg_out = QtGui.QPixmap(QtImg).scaled(
                self.ui_3.label_8.width(), self.ui_3.label_8.height())
            self.ui_3.label_8.setPixmap(jpg_out)

            """处理后"""
            pred_idx, confs, shrink_process = demo.process_frame(img)  # 返回的numppy格式,正好是opencv(RGB)
            # 这是必须的处理，否则画面会变得不一样
            shrink_process2 = cv2.cvtColor(shrink_process, cv2.COLOR_BGR2RGB)
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg_process = QtGui.QImage(shrink_process2.data,
                                         shrink_process2.shape[1],
                                         shrink_process2.shape[0],
                                         shrink_process2.shape[1] * 3,
                                         QtGui.QImage.Format_RGB888)

            # 修改图片的高宽
            jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
                self.ui_3.label_9.width(), self.ui_3.label_9.height())
            # 在控件上放置图片
            self.ui_3.label_9.setPixmap(jpg_out_process)
            for i in range(5):
                pred_class = self.labels[pred_idx[i]]  # 类别
                text = str(pred_class) + str('  ') + str(confs[i])  # 打印出来的格式
                self.ui_3.listWidget.addItem(text)

        else:
            """处理前"""
            self.ui_3.listWidget.clear()
            """对返回图片进行处理"""
            shrink = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg = QtGui.QImage(shrink.data,
                                 shrink.shape[1],
                                 shrink.shape[0],
                                 shrink.shape[1] * 3,
                                 QtGui.QImage.Format_RGB888)
            # print(QtImg)
            # 修改图片的高宽
            jpg_out = QtGui.QPixmap(QtImg).scaled(
                self.ui_3.label_8.width(), self.ui_3.label_8.height())
            self.ui_3.label_8.setPixmap(jpg_out)

            """处理后"""
            cv2.imwrite('temp.png', img)
            shrink_process, pred_idx, confs = demo.cam('temp.png')  # 返回的numppy格式,正好是opencv(RGB)
            # 这是必须的处理，否则画面会变得不一样
            shrink_process2 = cv2.cvtColor(shrink_process, cv2.COLOR_BGR2RGB)
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg_process = QtGui.QImage(shrink_process2.data,
                                         shrink_process2.shape[1],
                                         shrink_process2.shape[0],
                                         shrink_process2.shape[1] * 3,
                                         QtGui.QImage.Format_RGB888)

            # 修改图片的高宽
            jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
                self.ui_3.label_9.width(), self.ui_3.label_9.height())
            # 在控件上放置图片
            self.ui_3.label_9.setPixmap(jpg_out_process)
            for i in range(5):
                pred_class = self.labels[pred_idx[i]]  # 类别
                text = str(pred_class) + str('  ') + str(confs[i])  # 打印出来的格式
                self.ui_3.listWidget.addItem(text)
            os.remove('temp.png')

    """----------视频检测-----------"""

    """  ------------------------摄像头------------------------------ """

    def video_button(self):
        if self.flag_camera == 0:
            self.cap_video = cv2.VideoCapture(0)  # 开启摄像头
            self.timer_camera.start()  # 定时器开始，50ms后，进入timeout,除非遇到stop()，等时间一过（50ms）就执行timeout相连接的方法
            self.flag_camera = 1
        else:
            self.messageDialog()

    def show_camera(self):
        # 清除原来的文字信息
        self.ui_3.listWidget.clear()
        ret, self.img = self.cap_video.read()  # read（）返回两个参数：1、是否读取成功（布尔值） 2、这一帧的图片
        if ret:
            # 若成功读取——就展示（调用下面这个方法）
            self.show_cv_img(self.img)

    def show_cv_img(self, img):
        if not self.ui_3.checkBox.isChecked():
            shrink = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg = QtGui.QImage(shrink.data,
                                 shrink.shape[1],
                                 shrink.shape[0],
                                 shrink.shape[1] * 3,
                                 QtGui.QImage.Format_RGB888)

            # 修改图片的高宽
            jpg_out = QtGui.QPixmap(QtImg).scaled(
                self.ui_3.label_8.width(), self.ui_3.label_8.height())

            # 在控件上放置图片
            self.ui_3.label_8.setPixmap(jpg_out)

            # 处理后
            pred_idx, confs, shrink_process = demo.process_frame(img)  # 返回的numppy格式,正好是opencv(RGB)
            # 这是必须的处理，否则画面会变得不一样
            shrink_process2 = cv2.cvtColor(shrink_process, cv2.COLOR_BGR2RGB)
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg_process = QtGui.QImage(shrink_process2.data,
                                         shrink_process2.shape[1],
                                         shrink_process2.shape[0],
                                         shrink_process2.shape[1] * 3,
                                         QtGui.QImage.Format_RGB888)

            # 修改图片的高宽
            jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
                self.ui_3.label_9.width(), self.ui_3.label_9.height())

            # 在控件上放置图片
            self.ui_3.label_9.setPixmap(jpg_out_process)

            # 用文字表示
            for i in range(5):
                pred_class = self.labels[pred_idx[i]]  # 类别
                text = str(pred_class) + str('  ') + str(confs[i])  # 打印出来的格式
                self.ui_3.listWidget.addItem(text)

        else:
            shrink = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg = QtGui.QImage(shrink.data,
                                 shrink.shape[1],
                                 shrink.shape[0],
                                 shrink.shape[1] * 3,
                                 QtGui.QImage.Format_RGB888)

            # 修改图片的高宽
            jpg_out = QtGui.QPixmap(QtImg).scaled(
                self.ui_3.label_8.width(), self.ui_3.label_8.height())

            # 在控件上放置图片
            self.ui_3.label_8.setPixmap(jpg_out)

            # 处理后
            cv2.imwrite('temp.png', img)
            shrink_process, pred_idx, confs = demo.cam('temp.png')  # 返回的numppy格式,正好是opencv(RGB)
            # 这是必须的处理，否则画面会变得不一样
            shrink_process2 = cv2.cvtColor(shrink_process, cv2.COLOR_BGR2RGB)
            # 图像处理 通用  针对RGB三通道图片显示
            QtImg_process = QtGui.QImage(shrink_process2.data,
                                         shrink_process2.shape[1],
                                         shrink_process2.shape[0],
                                         shrink_process2.shape[1] * 3,
                                         QtGui.QImage.Format_RGB888)

            # 修改图片的高宽
            jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
                self.ui_3.label_9.width(), self.ui_3.label_9.height())

            # 在控件上放置图片
            self.ui_3.label_9.setPixmap(jpg_out_process)

            # 用文字表示
            for i in range(5):
                pred_class = self.labels[pred_idx[i]]  # 类别
                text = str(pred_class) + str('  ') + str(confs[i])  # 打印出来的格式
                self.ui_3.listWidget.addItem(text)

    """  ------------------------摄像头------------------------------ """

    def stoptest(self):
        if self.flag_img == 1:
            self.flag_img = 0
            self.ui_3.label_8.clear()
            self.ui_3.label_9.clear()
            self.ui_3.listWidget.clear()
        if self.flag_shipin == 1:
            self.flag_shipin = 0
            self.ui_3.label_8.clear()
            self.ui_3.label_9.clear()
            self.ui_3.listWidget.clear()
            self.timer_shipin.stop()
            self.cap_shipin.release()  # 释放资源
        if self.flag_camera == 1:
            self.flag_camera = 0
            self.ui_3.label_8.clear()
            self.ui_3.label_9.clear()
            self.ui_3.listWidget.clear()
            self.timer_camera.stop()
            self.cap_video.release()

    """-----------------保存结果-------------"""

    def save(self):
        img_path, ok = QInputDialog.getText(self, "图片名", "输入图片保存文件名")
        if img_path and ok:
            save_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), img_path + '.jpg')
            self.ui_3.label_9.pixmap().save(save_path)
            print(save_path)
            self.save_messageDialog(save_path)


# 窗口四
class my_window4(QMainWindow):

    def __init__(self):
        super(my_window4, self).__init__()
        self.ui_4 = Ui_Window_4()
        self.ui_4.setupUi(self)
        # 将边框隐藏起来
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.shadow = QtWidgets.QGraphicsDropShadowEffect(self)
        self.shadow.setOffset(0, 0)
        self.shadow.setBlurRadius(18)
        self.shadow.setColor(QtCore.Qt.black)
        self.ui_4.frame.setGraphicsEffect(self.shadow)
        # 标志
        self.image_flag = 0
        self.camera_flag = 0
        # 将槽函数与信号连接起来
        self.ui_4.pushButton_back.clicked.connect(self.back)
        self.ui_4.pushButton_min.clicked.connect(self.on_pushButton_min)
        self.ui_4.pushButton_close.clicked.connect(self.on_pushButton_close)
        self.ui_4.pushButton_wenjian.clicked.connect(self.showimage)
        self.ui_4.iouSpinBox.valueChanged.connect(self.iou_splider_change)  # 滑块的connect
        self.ui_4.iouSlider.valueChanged.connect(self.iou_spinbox_change)  # 微调框的connect
        self.ui_4.stopButton.clicked.connect(self.clear_label)
        self.ui_4.pushButton_wenjian.clicked.connect(self.showimage)
        self.ui_4.pushButton_shiping.clicked.connect(self.shipin_button)
        # 要实时显示显示，就要定时器
        timer = QTimer(self)  # 定义一个定时器对象
        timer.timeout.connect(self.showtime)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

        self.timer_camera = QtCore.QTimer()  # 创建一个定时器对象
        self.timer_camera.timeout.connect(self.show_camera)  # 定时器对象.timeout指的是当start()结束时，所调用的函数（进行操作）
        self.ui_4.pushButton_shexiangtou.clicked.connect(self.video_button)  # 摄像头按钮绑定video_button这个槽函数

        self.flag_shipin = 0
        self.timer_shipin = QTimer()
        self.timer_shipin.timeout.connect(self.show_shipin_img)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()  # 获取鼠标相对窗口的位置
            event.accept()
            self.setCursor(QCursor(Qt.OpenHandCursor))  # 更改鼠标图标

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_flag:
            self.move(QMouseEvent.globalPos() - self.m_Position)  # 更改窗口位置
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False
        self.setCursor(QCursor(Qt.ArrowCursor))

    def showtime(self):
        # 展示系统时间
        datetime = QDateTime.currentDateTime()  # 获取系统时间
        text = datetime.toString()
        self.ui_4.label_time.setText("   " + text)  # 在空间上展示

    def back(self):
        self.hide()
        win1.show()

    def on_pushButton_min(self):
        # 最小化
        self.showMinimized()

    def on_pushButton_close(self):
        # 关闭程序
        self.close()

    def iou_splider_change(self):
        self.ui_4.iouSpinBox.setValue(self.ui_4.iouSlider.value())

    def iou_spinbox_change(self):
        self.ui_4.iouSlider.setValue(self.ui_4.iouSpinBox.value())

    def showimage(self):

        if self.image_flag == 0:
            imgNamepath, imgType = QFileDialog.getOpenFileName(self, "选择图片",
                                                               "",
                                                               "*.jpg;;*.png;;All Files(*)")
            print(imgType)
            print(imgNamepath)
            if len(imgNamepath) > 0:
                self.image_flag = 1
                # 先保证是否能打开
                # 通过文件路径获取图片文件，并设置图片长宽为label控件的长、宽
                my_img = QPixmap(imgNamepath).scaled(self.ui_4.label_yunalai.width(), self.ui_4.label_yunalai.height())
                print(imgNamepath)
                # 在label控件上显示选择的图片
                self.ui_4.label_yunalai.setPixmap(my_img)

                """开始处理"""
                # self.window4_image_thread=New_Thread1(imgNamepath)  # 实例化一个线程
                # # 将线程thread的信号finishSignal和UI主线程中的槽函数Change进行连接
                # self.window4_image_thread.finishSignal.connect(self.showimage_process)
                # # 启动线程，执行线程类中run函数
                # self.window4_image_thread.start()
                opt = detect.parse_opt(imgNamepath)
                shrink_process, s = detect.main(opt)
                shrink_process2 = cv2.cvtColor(shrink_process, cv2.COLOR_BGR2RGB)
                QtImg_process = QtGui.QImage(shrink_process2.data,
                                             shrink_process2.shape[1],
                                             shrink_process2.shape[0],
                                             shrink_process2.shape[1] * 3,
                                             QtGui.QImage.Format_RGB888)
                # 修改图片的高宽,适应控件的大小
                jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
                    self.ui_4.label_chulihou.width(), self.ui_4.label_chulihou.height())

                # 在控件上放置图片
                self.ui_4.label_chulihou.setPixmap(jpg_out_process)

    def video_button(self):
        if self.camera_flag == 0:
            self.cap_video = cv2.VideoCapture(0)  # 开启摄像头
            # print('pp')
            self.timer_camera.start()  # 定时器开始，50ms后，进入timeout,除非遇到stop()，等时间一过（50ms）就执行timeout相连接的方法
        self.camera_flag = 1

    def show_camera(self):
        ret, img = self.cap_video.read()  # read（）返回两个参数：1、是否读取成功（布尔值） 2、这一帧的图片
        print(ret)
        shrink = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)  # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的
        # 图像处理 通用  针对RGB三通道图片显示
        QtImg = QtGui.QImage(shrink.data,
                             shrink.shape[1],
                             shrink.shape[0],
                             shrink.shape[1] * 3,
                             QtGui.QImage.Format_RGB888)

        # 修改图片的高宽
        jpg_out = QtGui.QPixmap(QtImg).scaled(
            self.ui_4.label_yunalai.width(), self.ui_4.label_yunalai.height())

        # 在控件上放置图片
        self.ui_4.label_yunalai.setPixmap(jpg_out)

        # 处理后
        cv2.imwrite('temp.png', img)
        # figure_save_path='temp_images'
        # os.makedirs(figure_save_path)  # 如果不存在目录figure_save_path，则创建
        # plt.savefig(os.path.join(figure_save_path, 'img'))  # 第一个是指存储路径，第二个是图片名字
        # print(figure_save_path)
        opt = detect.parse_opt(r'temp.png')  # 解析传入的参数
        shrink_process, s = detect.main(opt)
        # 这是必须的处理，否则画面会变得不一样
        shrink_process2 = cv2.cvtColor(shrink_process, cv2.COLOR_BGR2RGB)
        # 图像处理 通用  针对RGB三通道图片显示
        QtImg_process = QtGui.QImage(shrink_process2.data,
                                     shrink_process2.shape[1],
                                     shrink_process2.shape[0],
                                     shrink_process2.shape[1] * 3,
                                     QtGui.QImage.Format_RGB888)

        # 修改图片的高宽
        jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
            self.ui_4.label_chulihou.width(), self.ui_4.label_chulihou.height())

        # 在控件上放置图片
        self.ui_4.label_chulihou.setPixmap(jpg_out_process)

    def shipin_button(self):
        if self.flag_shipin == 0:
            # 弹出一个文件选择框，第一个返回值imgName记录选中的文件路径+文件名，第二个返回值imgType记录文件的类型
            # QFileDialog就是系统对话框的那个类第一个参数是上下文，第二个参数是弹框的名字，第三个参数是默认打开的路径，第四个参数是需要的格式
            shipinNamepath, shipinType = QFileDialog.getOpenFileName(self, "选择视频",
                                                                     "",
                                                                     "*.mp4;;*.avi;;All Files(*)")
            print(shipinType)

            # 得到文件后缀名  需要根据情况进行修改
            suffix = shipinNamepath.split("/")[-1][shipinNamepath.split("/")[-1].index(".") + 1:]
            print(shipinNamepath, suffix)

            if suffix == "mp4" or suffix == "avi":
                self.cap_shipin = cv2.VideoCapture(shipinNamepath)  # 0表示摄像头，返回值就是整个视屏资源
                self.timer_shipin.start()
            self.flag_shipin = 1

    def show_shipin_img(self):
        ret, self.shipin_img = self.cap_shipin.read()  # 读取每一帧图片，read（）返回两个参数：1、是否读取成功（布尔值） 2、这一帧的图片
        """处理前"""

        """对返回图片进行处理"""
        shrink = cv2.cvtColor(self.shipin_img, cv2.COLOR_BGR2RGB)  # 注意：视频读取的是BGR格式，需先转换为RGB格式，返回值就是转换后的
        # 图像处理 通用  针对RGB三通道图片显示
        QtImg = QtGui.QImage(shrink.data,
                             shrink.shape[1],
                             shrink.shape[0],
                             shrink.shape[1] * 3,
                             QtGui.QImage.Format_RGB888)
        # print(QtImg)
        # 修改图片的高宽
        jpg_out = QtGui.QPixmap(QtImg).scaled(
            self.ui_4.label_yunalai.width(), self.ui_4.label_yunalai.height())
        self.ui_4.label_yunalai.setPixmap(jpg_out)

        """处理后"""
        cv2.imwrite('temp.png', self.shipin_img)
        opt = detect.parse_opt(r'temp.png')  # 解析传入的参数
        shrink_process, s = detect.main(opt)
        # 这是必须的处理，否则画面会变得不一样
        shrink_process2 = cv2.cvtColor(shrink_process, cv2.COLOR_BGR2RGB)
        # 图像处理 通用  针对RGB三通道图片显示
        QtImg_process = QtGui.QImage(shrink_process2.data,
                                     shrink_process2.shape[1],
                                     shrink_process2.shape[0],
                                     shrink_process2.shape[1] * 3,
                                     QtGui.QImage.Format_RGB888)

        # 修改图片的高宽
        jpg_out_process = QtGui.QPixmap(QtImg_process).scaled(
            self.ui_4.label_chulihou.width(), self.ui_4.label_chulihou.height())

        # 在控件上放置图片
        self.ui_4.label_chulihou.setPixmap(jpg_out_process)

    def clear_label(self):
        if self.image_flag == 1:
            self.image_flag = 0
        if self.camera_flag == 1:
            self.timer_camera.stop()  # 如果再点一次摄像头图案（按钮），则此时flag=1，那么就是stop()，定时器停止
            self.camera_flag = 0
            self.cap_video.release()  # 释放资源
        if self.flag_shipin == 1:
            self.timer_shipin.stop()
            self.flag_shipin = 0
        self.ui_4.label_yunalai.clear()
        self.ui_4.label_chulihou.clear()


# 窗口五
class my_window5(QMainWindow):
    def __init__(self):
        super(my_window5, self).__init__()
        self.ui_5 = Ui_Window_5()
        self.ui_5.setupUi(self)
        # 将边框隐藏起来
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.shadow = QtWidgets.QGraphicsDropShadowEffect(self)
        self.shadow.setOffset(0, 0)
        self.shadow.setBlurRadius(18)
        self.shadow.setColor(QtCore.Qt.black)
        self.ui_5.frame.setGraphicsEffect(self.shadow)

        # 要实时显示显示，就要定时器
        timer = QTimer(self)  # 定义一个定时器对象
        timer.timeout.connect(self.showtime)  # timeout与showtime连接(当start()规定的时间结束后,就执行showtime)
        timer.start()  # 不写时间就默认为0,则相当于一直在timeout(),就是执行showtime()方法

        # 按钮与事件连接
        self.ui_5.pushButton.clicked.connect(self.back)
        self.ui_5.pushButton_4.clicked.connect(self.duqu)
        self.ui_5.pushButton_5.clicked.connect(self.daochu)
        self.ui_5.pushButton_2.clicked.connect(self.on_pushButton_min)
        self.ui_5.pushButton_3.clicked.connect(self.on_pushButton_close)

    """无边框控制移动"""

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()  # 获取鼠标相对窗口的位置
            event.accept()
            self.setCursor(QCursor(Qt.OpenHandCursor))  # 更改鼠标图标

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_flag:
            self.move(QMouseEvent.globalPos() - self.m_Position)  # 更改窗口位置
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False
        self.setCursor(QCursor(Qt.ArrowCursor))

    def showtime(self):
        # 展示系统时间
        datetime = QDateTime.currentDateTime()  # 获取系统时间
        text = datetime.toString()
        self.ui_5.label_2.setText("   " + text)  # 在空间上展示

    def on_pushButton_min(self):
        # 最小化
        self.showMinimized()

    def on_pushButton_close(self):
        # 关闭程序
        self.close()

    def back(self):
        self.hide()
        win1.show()

    def duqu(self):
        ###获取路径===================================================================

        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls , *.csv)')
        #  QFileDialog.getOpenFileName为打开文件，返回两个参数：1、文件名称（绝对路径）  2、文件类型（格式）
        # print(openfile_name)
        global path_openfile_name  # openfile_name分为两部分：1、文件名称（绝对路径）  2、文件类型（格式）

        ###获取路径====================================================================

        path_openfile_name = openfile_name[0]  # 得到路径
        # print(len(path_openfile_name))
        ###===========读取表格，转换表格，===========================================
        if len(path_openfile_name) > 0:
            # 先保证打开选择文件后，是否选中
            global input_table
            input_table = pd.read_excel(path_openfile_name)  # 用panda读取excel文件（参数为文件路径），返回值就是整个excel文件（还包括了行列数）
            # print(input_table)

            # 这里分别是行数和列数
            input_table_rows = input_table.shape[0]
            input_table_colunms = input_table.shape[1]  # input_table.shape为行列数（形式为列表）
            # print(input_table.shape)
            # print(input_table_rows)
            # print(input_table_colunms)

            input_table_header = input_table.columns.values.tolist()  # 表头  标题
            print(input_table_header)
            ###===========读取表格，转换表格，============================================

            ###======================给tablewidget设置行列表头============================
            # 将行数和列数传入，把tablewidget的结构构造好
            self.ui_5.tableWidget.setColumnCount(input_table_colunms)
            self.ui_5.tableWidget.setRowCount(input_table_rows)
            self.ui_5.tableWidget.setHorizontalHeaderLabels(input_table_header)

            ###======================给tablewidget设置行列表头============================

            ###================遍历表格每个元素，同时添加到tablewidget中========================
            for i in range(input_table_rows):
                input_table_rows_values = input_table.iloc[[i]]
                # print(input_table_rows_values)
                input_table_rows_values_array = np.array(input_table_rows_values)
                input_table_rows_values_list = input_table_rows_values_array.tolist()[0]
                # print(input_table_rows_values_list)
                for j in range(input_table_colunms):
                    input_table_items_list = input_table_rows_values_list[j]
                    # print(input_table_items_list)
                    # print(type(input_table_items_list))

                    ###==============将遍历的元素添加到tablewidget中并显示=======================

                    input_table_items = str(input_table_items_list)
                    newItem = QTableWidgetItem(input_table_items)
                    self.ui_5.tableWidget.setItem(i, j, newItem)

                    ###================遍历表格每个元素，同时添加到tablewidget中========================

    def daochu(self):
        # 参数：类本身、窗口名称、自动打开的初始路径（不设置就从本文件所在位置）、供选择的文件类型
        filename, _ = QFileDialog.getSaveFileName(self, "选择文件保存路径", "", "Excel files (*.xlsx)")
        # 返回两个值  文件地址、文件类型
        if filename:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active  # 激活
            # 保存表头
            for column in range(self.ui_5.tableWidget.columnCount()):
                header = self.ui_5.tableWidget.horizontalHeaderItem(column)
                worksheet.cell(row=1, column=column + 1, value=str(header.text()))
            # 保存数据
            for row in range(self.ui_5.tableWidget.rowCount()):
                for column in range(self.ui_5.tableWidget.columnCount()):
                    item = self.ui_5.tableWidget.item(row, column)
                    if item is not None:
                        worksheet.cell(row=row + 2, column=column + 1, value=str(item.text()))
            workbook.save(filename)
            # 小弹窗，提示保存成功
            QMessageBox.information(self, "提示", "数据文件保存成功!", QMessageBox.Ok)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    # 直接声明对象实例
    win1 = my_window1()
    win2 = my_window2()
    win3 = my_window3()
    win4 = my_window4()
    win5 = my_window5()
    win1.center()
    win1.show()
    sys.exit(app.exec_())  # 不间断的运行程序


